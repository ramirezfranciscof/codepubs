#!/usr/bin/python3
import matplotlib.pyplot as plt
import openpyxl
import math

################################################################################
# FUNCIONES
#------------------------------------------------------------------------------#

def read_notes( filename ):

    wb_src = openpyxl.load_workbook( filename )
    wb_sht = wb_src.get_active_sheet()
    datacol = wb_sht['A']

    datavals = []
    datanone = []

    for datacel in datacol:
        try:
            datavals.append( float(datacel.value) )
        except ValueError:
            datanone.append( datacel.value )


    return datavals


#------------------------------------------------------------------------------#

def setup_binlim( bins_min, bins_max, bins_num ):

    bins_lim0 = bins_min
    bins_size = (bins_max-bins_min)/bins_num

    bins_lims = [ ]

    for abin in range( bins_num ):
        bins_ulduo = ( bins_lim0, bins_lim0 + bins_size )
        bins_lims.append( bins_ulduo )
        bins_lim0 = bins_lim0 + bins_size

    return bins_lims

#------------------------------------------------------------------------------#

def populate_bins( datalist, binlims ):

    mysize  = len( binlims )
    binvals = [0] * mysize

    for dataval in datalist:
        for kbin in range( mysize ):

            mymin = binlims[kbin][0]
            mymax = binlims[kbin][1]

            if ( mymin < dataval <= mymax):
                binvals[kbin] = binvals[kbin] + 1

    return binvals

#------------------------------------------------------------------------------#

def write_bins( workbook, binlims, binvals ):

    mysheet = workbook.create_sheet( title='Statistics_0' )
    
    mysheet.cell(row=1, column=1).value = "bin min"
    mysheet.cell(row=1, column=2).value = "bin max"
    mysheet.cell(row=1, column=3).value = "bin val"
    
    mysize = len( binvals )
    for kbin in range( mysize ):
        mysheet.cell(row=2+kbin, column=1).value = binlims[kbin][0]
        mysheet.cell(row=2+kbin, column=2).value = binlims[kbin][1]
        mysheet.cell(row=2+kbin, column=3).value = binvals[kbin]
        
    return mysheet

#------------------------------------------------------------------------------#

def write_stat( workbook, datalist ):

    mysheet = workbook.create_sheet( title='Statistics_1' )

#   Calc mean
    mean = 0
    for dataval in datalist:
        mean = mean + dataval / len(datalist)

#   Calc stdev
    stdev = 0
    for dataval in datalist:
        addval = ( dataval - mean )
        stdev = stdev + addval * addval
    stdev = stdev / ( len(datalist) - 1 )
    stdev = math.sqrt( stdev )

    mysheet.cell(row=1, column=1).value = "Mean Note:"
    mysheet.cell(row=1, column=2).value = mean
    mysheet.cell(row=2, column=1).value = "Std Dev:"
    mysheet.cell(row=2, column=2).value = stdev

        
    return mysheet

#------------------------------------------------------------------------------#

def plots_bins( binlims, binvals ):

    bincenter = []
    mywidth = ( (binlims[0][1]-binlims[0][0])/2 )
    mybase  = mywidth / 2

    for newlim in binlims:
        bincenter.append( (newlim[0]+newlim[1])/2 - mybase )

    plt.bar( bincenter, binvals, width=mywidth )
    plt.show()
    return True

################################################################################
# PROCEDIMIENTOS
#------------------------------------------------------------------------------#

notas = read_notes( 'notasqi.xlsx' )

bins_num = 10
bins_limits = setup_binlim( 0, 100, bins_num )
bins_values = populate_bins( notas, bins_limits )

wb = openpyxl.Workbook()
wb.remove_sheet( wb.get_active_sheet() )

write_bins( wb, bins_limits, bins_values )
plots_bins( bins_limits, bins_values )
write_stat( wb, notas )

wb.save('notasqi_stats.xlsx')

################################################################################
